import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import fitz  # PyMuPDF
import os
from PIL import Image, ImageTk
import numpy as np
import pandas as pd
from io import BytesIO
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill
import string
import math
import tensorflow as tf
from tensorflow.keras.preprocessing import image
from openpyxl import load_workbook

def generate_letter_list(n):
    letters = string.ascii_uppercase
    start_index = letters.index('D')
    result = [letters[(start_index + i) % 26] for i in range(n)]
    return result

def rgb_to_hex(rgb):
    return "{:02x}{:02x}{:02x}".format(rgb[0], rgb[1], rgb[2])

def apply_color_to_cells(ws, rgb_columns):
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter in rgb_columns:
                if isinstance(cell.value, str) and cell.value.startswith('(') and cell.value.endswith(')'):
                    try:
                        rgb = tuple(map(int, cell.value.strip('()').split(',')))
                        hex_color = rgb_to_hex(rgb)
                        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        cell.fill = fill
                    except ValueError:
                        continue

def select_pdf_file():
    """
    选择多个PDF文件
    :return: 选择的多个文件路径
    """
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    if file_paths:
        pdf_folder_entry.delete(0, tk.END)
        pdf_folder_entry.insert(0, ", ".join(file_paths))

def select_output_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        output_folder_entry.delete(0, tk.END)
        output_folder_entry.insert(0, folder_path)

def select_csv_file():
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    if file_path:
        csv_file_entry.delete(0, tk.END)
        csv_file_entry.insert(0, file_path)

def select_model():
    file_path = filedialog.askopenfilename(defaultextension=".h5")
    if file_path:
        model_entry.delete(0, tk.END)
        model_entry.insert(0, file_path)

def extract_images_from_pdf(file_paths, output_image_folder, min_width=50, min_height=50):
    image_data = []
    print(file_paths)
    for pdf_file_path in file_paths:
        document = fitz.open(pdf_file_path)
        for page_index in range(len(document)):
            page = document.load_page(page_index)
            page_width, page_height = page.rect.width, page.rect.height
            image_list = page.get_images(full=True)
            for img_index, img in enumerate(image_list):
                xref = img[0]
                base_image = document.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                image_filename = f"{os.path.splitext(os.path.basename(pdf_file_path))[0]}_page_{page_index + 1}_img_{img_index + 1}.{image_ext}"
                image_path = os.path.join(output_image_folder, image_filename)

                image = Image.open(BytesIO(image_bytes))
                orig_width, orig_height = image.size
                if orig_width < min_width or orig_height < min_height:
                    continue
                rect = page.get_image_bbox(img)
                displayed_width = rect.width
                displayed_height = rect.height
                ratio = displayed_width * displayed_height / (page_width * page_height)
                if ratio > 1:
                    ratio = 1
                pixels = np.array(image)
                std = pixels.std(axis=0)
                if std.mean() <= 30:
                    continue

                with open(image_path, "wb") as image_file:
                    image_file.write(image_bytes)
                image_data.append({
                    "pdf_file": os.path.basename(pdf_file_path),
                    "image_file": image_filename,
                    "page_number": page_index + 1,
                    "image_path": image_path,
                    "ratio": ratio
                })
    return image_data

def color_distance(c1, c2):
    r1, g1, b1 = c1
    r2, g2, b2 = c2
    return ((r1 - r2) ** 2 + (g1 - g2) ** 2 + (b1 - b2) ** 2) ** 0.5

def color_statistics(image_path, top_n_colors, similarity_threshold=30):
    image = Image.open(image_path)
    image = image.convert('RGB')
    pixels = np.array(image)
    pixels = pixels.reshape((-1, 3))
    counter = Counter(tuple(pixel) for pixel in pixels)
    most_common_colors = counter.most_common()

    filtered_colors = []
    for color, count in most_common_colors:
        if len(filtered_colors) >= top_n_colors:
            break
        is_similar = False
        for f_color, _ in filtered_colors:
            if color_distance(color, f_color) < similarity_threshold:
                is_similar = True
                break
        if not is_similar:
            filtered_colors.append((color, count))

    return filtered_colors[:top_n_colors]

def average_rgb(colors):
    total = [0, 0, 0]
    count = 0
    for value in colors.values():
        for i in range(len(value)):
            total[i] += value[i]
        count += 1
    average = [total[i] // count for i in range(len(total))]
    return average

def process_selected_images(image_data, model_path, output_image_folder):
    output_csv_path = csv_file_entry.get()
    top_n_colors = int(top_n_colors_entry.get())
    similarity_threshold = float(similarity_threshold_entry.get())
    results = []
    total_images = len(image_data)

    for idx, data in enumerate(image_data):
        image_path = data["image_path"]
        color_counts = color_statistics(image_path, top_n_colors, similarity_threshold)
        color_info = {f"color_{i + 1}": color[0] for i, color in enumerate(color_counts)}
        color_mean = average_rgb(color_info)
        color_info.update({f"color_{i + 1}_count": color[1] for i, color in enumerate(color_counts)})

        result = {
            "pdf_file": data["pdf_file"],
            "image_file": data["image_file"],
            "page_number": data["page_number"],
            "ratio": data['ratio'],
            'color_mean': color_mean
        }
        result.update(color_info)
        results.append(result)

    df = pd.DataFrame(results)
    df.to_csv(output_csv_path, index=False)

    # 进行预测
    os.environ["CUDA_VISIBLE_DEVICES"] = "cpu"
    model = tf.keras.models.load_model(model_path)

    image_files = df['image_file'].tolist()
    labels_dict = {0: '中性图片', 1: '正性图片', 2: '负性图片'}
    pred_labels = []

    for image_file in image_files:
        img_path = os.path.join(output_image_folder, image_file)
        if not os.path.exists(img_path):
            pred_labels.append('文件不存在')
            continue

        img = image.load_img(img_path, target_size=(224, 224))
        img_array = image.img_to_array(img)
        img_array = np.expand_dims(img_array, axis=0)
        predictions = model.predict(img_array)
        predicted_label = np.argmax(predictions, axis=1)
        pred_labels.append(labels_dict[predicted_label[0]])

    df['predicted_label'] = pred_labels
    df.to_csv(output_csv_path, index=False)
    output_excel_path = output_csv_path.replace(".csv", ".xlsx")
    df.to_excel(output_excel_path, index=False, engine='openpyxl')

    wb = openpyxl.load_workbook(output_excel_path)
    ws = wb.active
    rgb_columns = generate_letter_list(top_n_colors)
    apply_color_to_cells(ws, rgb_columns)
    wb.save(output_excel_path)

    messagebox.showinfo("完成", f"结果已保存到 {output_excel_path}")
def show_image_selection_window(image_data, output_image_folder):
    """
    显示图片选择窗口
    :param image_data: 包含所有图片路径和其他信息
    :param output_image_folder: 输出文件夹路径
    :return: None
    """
    selection_window = tk.Toplevel(app)
    selection_window.title("选择要保留的图片")
    selection_window.geometry("800x600")

    # 打开图片所在文件夹
    images_in_folder = [data["image_path"] for data in image_data]
    selected_files = filedialog.askopenfilenames(
        initialdir=output_image_folder,
        title="选择需要保留的图片",
        filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")]
    )

    # 将用户选择的图片路径存储为集合
    selected_files_set = set(selected_files)

    # 删除未选择的图片
    for image_path in images_in_folder:
        if image_path not in selected_files_set:
            if os.path.exists(image_path):
                os.remove(image_path)

    # 关闭选择窗口
    selection_window.destroy()
    return list(selected_files_set)
def process_pdfs():
    pdf_file_path = pdf_folder_entry.get().split(", ")
    output_image_folder = output_folder_entry.get()
    output_csv_path = csv_file_entry.get()
    model_path = model_entry.get()

    if not pdf_file_path or not output_image_folder or not output_csv_path:
        messagebox.showerror("错误", "请填写所有输入框")
        return

    if not os.path.exists(output_image_folder):
        os.makedirs(output_image_folder)

    image_data = extract_images_from_pdf(pdf_file_path, output_image_folder)
    show_image_selection_window(image_data, output_image_folder)
    process_selected_images(image_data, model_path, output_image_folder)

app = tk.Tk()
app.title("PDF处理程序")

pdf_folder_label = tk.Label(app, text="选择PDF文件:")
pdf_folder_label.pack()
pdf_folder_entry = tk.Entry(app, width=50)
pdf_folder_entry.pack()
pdf_folder_button = tk.Button(app, text="选择", command=select_pdf_file)
pdf_folder_button.pack()

output_folder_label = tk.Label(app, text="选择输出文件夹:")
output_folder_label.pack()
output_folder_entry = tk.Entry(app, width=50)
output_folder_entry.pack()
output_folder_button = tk.Button(app, text="选择", command=select_output_folder)
output_folder_button.pack()

csv_file_label = tk.Label(app, text="选择输出CSV文件:")
csv_file_label.pack()
csv_file_entry = tk.Entry(app, width=50)
csv_file_entry.pack()
csv_file_button = tk.Button(app, text="选择", command=select_csv_file)
csv_file_button.pack()

model_label = tk.Label(app, text="选择模型文件:")
model_label.pack()
model_entry = tk.Entry(app, width=50)
model_entry.pack()
model_button = tk.Button(app, text="选择", command=select_model)
model_button.pack()

top_n_colors_label = tk.Label(app, text="选择返回的颜色数量:")
top_n_colors_label.pack()
top_n_colors_entry = tk.Entry(app, width=50)
top_n_colors_entry.insert(0, "5")
top_n_colors_entry.pack()

similarity_threshold_label = tk.Label(app, text="设置颜色相似度阈值 (0-255):")
similarity_threshold_label.pack()
similarity_threshold_entry = tk.Entry(app, width=50)
similarity_threshold_entry.insert(0, "30")
similarity_threshold_entry.pack()

process_button = tk.Button(app, text="处理PDF", command=process_pdfs)
process_button.pack()

app.mainloop()
