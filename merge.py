import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import fitz  # PyMuPDF
import os
from PIL import Image, ImageTk
import numpy as np
import pandas as pd
from io import BytesIO
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill
import string
import math
import tensorflow as tf
from tensorflow.keras.preprocessing import image
from openpyxl import load_workbook
def generate_letter_list(n):
    # 获取所有大写字母的字符串
    letters = string.ascii_uppercase
    # 找到'D'的索引位置
    start_index = letters.index('D')
    # 生成从'D'开始的n个字母的列表
    result = [letters[(start_index + i) % 26] for i in range(n)]
    return result
def rgb_to_hex(rgb):
    return "{:02x}{:02x}{:02x}".format(rgb[0], rgb[1], rgb[2])



def apply_color_to_cells(ws, rgb_columns):
    """
    将csv文件中的rgb元组转换成可视颜色
    :param ws:
    :param rgb_columns:
    :return:
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter in rgb_columns:
                if isinstance(cell.value, str) and cell.value.startswith('(') and cell.value.endswith(')'):
                    try:
                        rgb = tuple(map(int, cell.value.strip('()').split(',')))
                        hex_color = rgb_to_hex(rgb)
                        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        cell.fill = fill
                    except ValueError:
                        continue

def select_pdf_files():
    """
    选择多个PDF文件
    :return: 选择的多个文件路径
    """
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    if file_paths:
        pdf_folder_entry.delete(0, tk.END)
        pdf_folder_entry.insert(0, ", ".join(file_paths))


def select_output_folder():
    """
    选择图片输出文件夹
    :return:
    """
    folder_path = filedialog.askdirectory()
    if folder_path:
        output_folder_entry.delete(0, tk.END)
        output_folder_entry.insert(0, folder_path)


def select_csv_file():
    """
    选择csv输出文件
    :return:
    """
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
    if file_path:
        csv_file_entry.delete(0, tk.END)
        csv_file_entry.insert(0, file_path)

def select_model():
    file_path = filedialog.askopenfilename(defaultextension=".h5")
    if file_path:
        model_entry.delete(0, tk.END)
        model_entry.insert(0, file_path)
def extract_images_from_pdf(file_paths, output_image_folder, min_width=50, min_height=50):
    image_data = []
    print(file_paths)
    for pdf_file_path in file_paths:
        document = fitz.open(pdf_file_path)
        for page_index in range(len(document)):
            page = document.load_page(page_index)
            page_width, page_height = page.rect.width, page.rect.height
            image_list = page.get_images(full=True)
            for img_index, img in enumerate(image_list):
                try:
                    xref = img[0]
                    base_image = document.extract_image(xref)
                    image_bytes = base_image["image"]
                    image_ext = base_image["ext"]
                    image_filename = f"{os.path.splitext(os.path.basename(pdf_file_path))[0]}_page_{page_index + 1}_img_{img_index + 1}.{image_ext}"
                    image_path = os.path.join(output_image_folder, image_filename)

                    image = Image.open(BytesIO(image_bytes))
                    orig_width, orig_height = image.size
                    if orig_width < min_width or orig_height < min_height:
                        continue
                    rect = page.get_image_bbox(img)
                    displayed_width = rect.width
                    displayed_height = rect.height
                    ratio = displayed_width * displayed_height / (page_width * page_height)
                    if ratio > 1:
                        ratio = 1
                    pixels = np.array(image)
                    std = pixels.std(axis=0)
                    if std.mean() <= 30:
                        continue

                    with open(image_path, "wb") as image_file:
                        image_file.write(image_bytes)
                    image_data.append({
                        "pdf_file": os.path.basename(pdf_file_path),
                        "image_file": image_filename,
                        "page_number": page_index + 1,
                        "image_path": image_path,
                        "ratio": ratio
                    })
                except Exception as e:
                    print(e)
    return image_data

def ColourDistance(rgb_1, rgb_2):
    """
    计算两个rgb颜色相似度
    :param rgb_1: 颜色1
    :param rgb_2: 颜色2
    :return: 两个rgb颜色相似度
    """
    R_1, G_1, B_1 = rgb_1
    R_2, G_2, B_2 = rgb_2
    rmean = (R_1 + R_2) / 2
    R = R_1 - R_2
    G = G_1 - G_2
    B = B_1 - B_2
    return math.sqrt((2 + rmean / 256) * (R ** 2) + 4 * (G ** 2) + (2 + (255 - rmean) / 256) * (B ** 2))

def color_distance(c1, c2):
    r1,g1,b1=c1
    r2,g2,b2=c2
    return ((r1-r2) ** 2+(g1-g2)**2+(b1-b2)**2) ** 0.5
# def color_statistics(image_path, top_n_colors):
#     image = Image.open(image_path)
#     image = image.convert('RGB')
#     pixels = np.array(image)
#     pixels = pixels.reshape((-1, 3))
#     counter = Counter(tuple(pixel) for pixel in pixels)
#     most_common_colors = counter.most_common(top_n_colors)
#     return most_common_colors
def color_statistics(image_path, top_n_colors, similarity_threshold=30):
    """
    将统计的所有rgb颜色作比较，颜色详尽的数量较少的颜色直接剔除
    :param image_path: 图像路径
    :param top_n_colors: 统计的前n个颜色
    :param similarity_threshold: 相似度阈值
    :return: 前top_n_colors颜色
    """
    image = Image.open(image_path)
    image = image.convert('RGB')
    pixels = np.array(image)
    pixels = pixels.reshape((-1, 3))
    counter = Counter(tuple(pixel) for pixel in pixels)
    most_common_colors = counter.most_common()

    # 去重后的颜色列表
    filtered_colors = []

    for color, count in most_common_colors:
        if len(filtered_colors) >= top_n_colors:
            break
        is_similar = False
        for f_color, _ in filtered_colors:
            if color_distance(color, f_color) < similarity_threshold:
                is_similar = True
                break
        if not is_similar:
            filtered_colors.append((color, count))

    return filtered_colors[:top_n_colors]

#
# def update_progress(progress):
#     """
#     进度条设置
#     :param progress:
#     :return:
#     """
#     progress_var.set(progress)
#     app.update_idletasks()


def show_image_selection_window(image_data, model_path, output_image_folder):
    """
    显示图片展示窗口
    :param image_data: 包含所有图片路径和其他信息
    :param model_path: 模型路径
    :param output_image_folder: 输出文件夹路径
    :return: None
    """
    selection_window = tk.Toplevel(app)
    selection_window.title("选择要保留的图片")
    selection_window.geometry("800x600")

    # 创建滚动框架和滚动条
    canvas = tk.Canvas(selection_window, bg="white")
    scrollbar = tk.Scrollbar(selection_window, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    # 绑定滚动事件来通过鼠标滚轮滚动
    def on_mouse_wheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    canvas.bind_all("<MouseWheel>", on_mouse_wheel)  # Windows上滚动事件
    canvas.bind_all("<Button-4>", on_mouse_wheel)    # Linux/Mac上的滚动事件
    canvas.bind_all("<Button-5>", on_mouse_wheel)    # Linux/Mac上的滚动事件

    # 配置canvas窗口和滚动条
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # 存储复选框状态的字典
    checkboxes = {}
    selected_images = []

    # 遍历并显示所有图片
    for data in image_data:
        image_path = data["image_path"]
        img = Image.open(image_path)
        img.thumbnail((150, 150))  # 缩小图片
        img = ImageTk.PhotoImage(img)
        frame = tk.Frame(scrollable_frame, relief=tk.RAISED, bd=2)
        frame.pack(pady=5, padx=5, fill="x")

        label = tk.Label(frame, image=img)
        label.image = img  # 防止图像被垃圾回收
        label.pack(side="left")

        var = tk.BooleanVar(value=True)  # 默认选中
        checkbutton = tk.Checkbutton(frame, text=image_path, variable=var)
        checkbutton.pack(side="left")
        checkboxes[image_path] = var

    # 点击完成按钮时的处理
    def on_done():
        # 获取所有选中的图片路径
        selected_images.clear()
        for image_path, var in checkboxes.items():
            if var.get():  # 如果勾选了该图片
                selected_images.append(image_path)
        select_image_data=image_data.copy()
        # 自动删除未选择的图片
        for data in image_data:
            if data["image_path"] not in selected_images:
                image_path = data["image_path"]
                if os.path.exists(image_path):
                    os.remove(image_path)
                # image_data.remove(data)
                select_image_data.remove(data)
        process_selected_images(select_image_data, model_path, output_image_folder)
        # 这里你可以进一步处理所选中的图片
        # 例如，进行模型预测等操作
        selection_window.destroy()

    # # 完成按钮
    # done_button = tk.Button(selection_window, text="完成", command=on_done)
    # done_button.pack(pady=10)
    #
    #
    #
    # def on_done():
    #     selected_image_data = [data for data in image_data if checkboxes[data["image_path"]].get()]
    #     for data in image_data:
    #         if not checkboxes[data["image_path"]].get():
    #             os.remove(data["image_path"])
    #     selection_window.destroy()
    #     process_selected_images(selected_image_data,model_path,output_image_folder)
    done_button = tk.Button(selection_window, text="完成", command=on_done)
    done_button.pack(pady=10)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")


def average_rgb(colors):
    total = [0, 0, 0]  # 假设每个元组有3个元素
    count = 0
    # print(colors)
    for value in colors.values():

        for i in range(len(value)):
            total[i] += value[i]
        count += 1

    # 计算平均值
    average = [total[i] // count for i in range(len(total))]

    return average

# colors={1:(134,221,1),2:(134,221,1),3:(134,221,1),4:(134,221,1)}
# ave=average_rgb(colors)
# print(ave)
def process_selected_images(image_data,model_path,output_image_folder):
    output_csv_path = csv_file_entry.get()
    top_n_colors = int(top_n_colors_entry.get())
    similarity_threshold = float(similarity_threshold_entry.get())
    results = []
    total_images = len(image_data)

    for idx, data in enumerate(image_data):
        image_path = data["image_path"]
        # update_progress((idx + 1) / total_images * 100)

        color_counts = color_statistics(image_path, top_n_colors,similarity_threshold)
        color_info = {f"color_{i + 1}": color[0] for i, color in enumerate(color_counts)}
        color_mean=average_rgb(color_info)
        color_info.update({f"color_{i + 1}_count": color[1] for i, color in enumerate(color_counts)})



        result = {
            "pdf_file": data["pdf_file"],
            "image_file": data["image_file"],
            "page_number": data["page_number"],
            "ratio":data['ratio'],
            'color_mean':color_mean
        }
        result.update(color_info)
        results.append(result)

    df = pd.DataFrame(results)
    # df.to_csv(output_csv_path, index=False)

    # 进行预测
    os.environ["CUDA_VISIBLE_DEVICES"] = "cpu"

    # 加载模型
    model = tf.keras.models.load_model(model_path)



    # 获取image_file列的所有图片文件名
    image_files = df['image_file'].tolist()

    # 定义标签字典
    labels_dict = {0: '中性图片', 1: '正性图片', 2: '负性图片'}

    pred_labels = []

    # 遍历所有图片
    for image_file in image_files:
        img_path = os.path.join(output_image_folder, image_file)

        # 检查图像文件是否存在
        if not os.path.exists(img_path):
            pred_labels.append('文件不存在')
            continue

        # 加载图片
        img = image.load_img(img_path, target_size=(224, 224))
        img_array = image.img_to_array(img)
        img_array = np.expand_dims(img_array, axis=0)  # 添加批次维度

        # 进行预测
        predictions = model.predict(img_array)
        predicted_label = np.argmax(predictions, axis=1)  # 获取预测标签
        pred_labels.append(labels_dict[predicted_label[0]])

    # 将预测标签写入新列
    df['predicted_label'] = pred_labels
    df.to_csv(output_csv_path, index=False)
    output_excel_path=output_csv_path.replace(".csv", ".xlsx")
    df.to_excel(output_excel_path, index=False, engine='openpyxl')

    # 打开Excel文件
    wb = openpyxl.load_workbook(output_excel_path)
    ws = wb.active
    rgb_columns=generate_letter_list(top_n_colors)
    # 应用颜色到指定列
    apply_color_to_cells(ws, rgb_columns)

    # 保存Excel文件
    wb.save(output_excel_path)
    messagebox.showinfo("完成", f"结果已保存到 {output_excel_path}")
    # update_progress(0)


def process_pdfs():
    pdf_file_paths = pdf_folder_entry.get().split(", ")
    output_image_folder = output_folder_entry.get()
    output_csv_path = csv_file_entry.get()
    model_path = model_entry.get()

    if not pdf_file_paths or not output_image_folder or not output_csv_path:
        messagebox.showerror("错误", "请填写所有输入框")
        return

    if not os.path.exists(output_image_folder):
        os.makedirs(output_image_folder)

    image_data = extract_images_from_pdf(pdf_file_paths,output_image_folder)
    show_image_selection_window(image_data, model_path, output_image_folder)
    # process_selected_images(image_data, model_path, output_image_folder)





app = tk.Tk()
app.title("PDF处理程序")

pdf_folder_label = tk.Label(app, text="PDF文件:")
pdf_folder_label.pack()
pdf_folder_entry = tk.Entry(app, width=50)
pdf_folder_entry.pack()
pdf_folder_button = tk.Button(app, text="选择", command=select_pdf_files)
pdf_folder_button.pack()

output_folder_label = tk.Label(app, text="输出图片文件夹:")
output_folder_label.pack()
output_folder_entry = tk.Entry(app, width=50)
output_folder_entry.pack()
output_folder_button = tk.Button(app, text="选择文件夹", command=select_output_folder)
output_folder_button.pack()

model_lable=tk.Label(app, text="选择模型文件 (.h5):")
model_lable.pack()
model_entry = tk.Entry(app, width=50)
model_entry.pack()
# model_entry.grid(row=0, column=1, padx=10, pady=10)
model_button = tk.Button(app, text="浏览...", command= select_model)
# model_button.grid(row=0, column=2, padx=10, pady=10)
model_button.pack()



csv_file_label = tk.Label(app, text="输出CSV文件:")
csv_file_label.pack()
csv_file_entry = tk.Entry(app, width=50)
csv_file_entry.pack()
csv_file_button = tk.Button(app, text="选择文件", command=select_csv_file)
csv_file_button.pack()

top_n_colors_label = tk.Label(app, text="统计前多少颜色:")
top_n_colors_label.pack()
top_n_colors_entry = tk.Entry(app, width=10)
top_n_colors_entry.pack()

similarity_threshold_label = tk.Label(app, text="设置颜色相似度阈值 (0-255):")
similarity_threshold_label.pack()
similarity_threshold_entry = tk.Entry(app, width=50)
similarity_threshold_entry.insert(0, "30")
similarity_threshold_entry.pack()

process_button = tk.Button(app, text="开始处理", command=process_pdfs)
process_button.pack()

app.mainloop()
