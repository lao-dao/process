import os
import tensorflow as tf
from tensorflow.keras.preprocessing import image
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def rgb_to_hex(rgb):
    return "{:02x}{:02x}{:02x}".format(rgb[0], rgb[1], rgb[2])
def apply_color_to_cells(ws, rgb_columns):
    """
    将csv文件中的rgb元组转换成可视颜色
    :param ws:
    :param rgb_columns:
    :return:
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter in rgb_columns:
                if isinstance(cell.value, str) and cell.value.startswith('(') and cell.value.endswith(')'):
                    try:
                        rgb = tuple(map(int, cell.value.strip('()').split(',')))
                        hex_color = rgb_to_hex(rgb)
                        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        cell.fill = fill
                    except ValueError:
                        continue
def predict_images_from_xlsx(model_path, image_folder, xlsx_path):
    os.environ["CUDA_VISIBLE_DEVICES"] = "cpu"

    # 加载模型
    model = tf.keras.models.load_model(model_path)

    # 读取xlsx文件
    df = pd.read_excel(xlsx_path)

    # 获取image_file列的所有图片文件名
    image_files = df['image_file'].tolist()

    # 定义标签字典
    labels_dict = {0: '中性图片', 1: '正性图片', 2: '负性图片'}

    pred_labels = []

    # 遍历所有图片
    for image_file in image_files:
        img_path = os.path.join(image_folder, image_file)

        # 检查图像文件是否存在
        if not os.path.exists(img_path):
            pred_labels.append('文件不存在')
            continue

        # 加载图片
        img = image.load_img(img_path, target_size=(224, 224))
        img_array = image.img_to_array(img)
        img_array = np.expand_dims(img_array, axis=0)  # 添加批次维度

        # 进行预测
        predictions = model.predict(img_array)
        predicted_label = np.argmax(predictions, axis=1)  # 获取预测标签
        pred_labels.append(labels_dict[predicted_label[0]])

    # 将预测标签写入新列
    df['predicted_label'] = pred_labels

    # 将结果保存回xlsx文件
    df.to_excel(xlsx_path, index=False)
    # 加载保存的xlsx文件以设置颜色
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 对每一行，根据color_n列的RGB值为单元格设置背景颜色
    color_columns = [col for col in df.columns if col.startswith('color_') and not col.endswith(('count','mean'))]

    for index, row in df.iterrows():
        for col in color_columns:
            rgb_value = row[col]
            # RGB值可能是字符串，例如 "255,255,255"
            r, g, b = map(int, rgb_value.strip('()').split(','))
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            ws.cell(row=index + 2, column=df.columns.get_loc(col) + 1).fill = fill

    # 保存包含颜色的新文件
    wb.save(xlsx_path)
    messagebox.showinfo("完成", f"预测结果已保存到: {xlsx_path}")


# predict_images_from_xlsx('bestmodel.h5', 'output', 'results-test.xlsx')

def select_file(entry):
    file_path = filedialog.askopenfilename()
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)


def select_folder(entry):
    folder_path = filedialog.askdirectory()
    if folder_path:
        entry.delete(0, tk.END)
        entry.insert(0, folder_path)


def run_prediction():
    model_path = model_entry.get()
    image_folder = folder_entry.get()
    xlsx_path = xlsx_entry.get()

    if not model_path or not image_folder or not xlsx_path:
        messagebox.showwarning("警告", "请确保所有路径都已选择！")
        return

    try:
        predict_images_from_xlsx(model_path, image_folder, xlsx_path)
    except Exception as e:
        messagebox.showerror("错误", f"预测过程中出现错误: {e}")


# 创建主窗口
root = tk.Tk()
root.title("图片分类预测")

# 模型文件选择
tk.Label(root, text="选择模型文件 (.h5):").grid(row=0, column=0, padx=10, pady=10)
model_entry = tk.Entry(root, width=50)
model_entry.grid(row=0, column=1, padx=10, pady=10)
model_button = tk.Button(root, text="浏览...", command=lambda: select_file(model_entry))
model_button.grid(row=0, column=2, padx=10, pady=10)

# 图片文件夹选择
tk.Label(root, text="选择图片文件夹:").grid(row=1, column=0, padx=10, pady=10)
folder_entry = tk.Entry(root, width=50)
folder_entry.grid(row=1, column=1, padx=10, pady=10)
folder_button = tk.Button(root, text="浏览...", command=lambda: select_folder(folder_entry))
folder_button.grid(row=1, column=2, padx=10, pady=10)

# XLSX文件选择
tk.Label(root, text="选择XLSX文件:").grid(row=2, column=0, padx=10, pady=10)
xlsx_entry = tk.Entry(root, width=50)
xlsx_entry.grid(row=2, column=1, padx=10, pady=10)
xlsx_button = tk.Button(root, text="浏览...", command=lambda: select_file(xlsx_entry))
xlsx_button.grid(row=2, column=2, padx=10, pady=10)

# 运行按钮
run_button = tk.Button(root, text="开始预测", command=run_prediction)
run_button.grid(row=3, column=0, columnspan=3, pady=20)

root.mainloop()
